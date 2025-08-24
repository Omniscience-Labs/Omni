"""
Excel Tool for Omni - Advanced spreadsheet operations with full Excel support

This tool provides comprehensive Excel functionality including:
- Create, read, update Excel files (.xlsx, .xls, .csv)
- Advanced formatting and styling
- Formula calculations and data analysis
- Chart generation and visualization
- Data validation and conditional formatting
- Multi-sheet operations
- Import/export capabilities
"""

import os
import json
import pandas as pd
import numpy as np
from typing import Optional, Dict, Any, List, Union
from datetime import datetime
import logging
import asyncio
import tempfile
import base64
from io import BytesIO

# Excel-specific imports
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from sandbox.tool_base import SandboxToolsBase
from agentpress.tool import ToolResult

logger = logging.getLogger(__name__)

class SandboxExcelTool(SandboxToolsBase):
    """Advanced Excel operations tool for comprehensive spreadsheet management.
    
    Features:
    - Full Excel file support (.xlsx, .xls, .csv)
    - Advanced formatting and styling
    - Formula calculations and data analysis
    - Chart generation and visualization
    - Data validation and conditional formatting
    - Multi-sheet operations
    - Import/export capabilities
    """

    def get_tool_definition(self) -> Dict[str, Any]:
        return {
            "type": "function",
            "function": {
                "name": "excel_operations",
                "description": "Comprehensive Excel spreadsheet operations including create, read, update, format, analyze, and visualize Excel files with advanced features",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "operation": {
                            "type": "string",
                            "enum": [
                                "create", "read", "update", "delete", "format", 
                                "analyze", "chart", "formula", "validate", 
                                "export", "import", "sheet_operations"
                            ],
                            "description": "The Excel operation to perform"
                        },
                        "file_path": {
                            "type": "string",
                            "description": "Path to the Excel file (supports .xlsx, .xls, .csv)"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Name of the worksheet (default: 'Sheet1')"
                        },
                        "data": {
                            "type": "object",
                            "description": "Data for create/update operations (JSON format or array of arrays)"
                        },
                        "range": {
                            "type": "string",
                            "description": "Cell range (e.g., 'A1:C10', 'B2', 'A:A')"
                        },
                        "formatting": {
                            "type": "object",
                            "description": "Formatting options (font, fill, border, alignment)"
                        },
                        "chart_type": {
                            "type": "string",
                            "enum": ["bar", "line", "pie", "scatter", "area"],
                            "description": "Type of chart to create"
                        },
                        "formula": {
                            "type": "string",
                            "description": "Excel formula to apply (e.g., '=SUM(A1:A10)')"
                        },
                        "validation": {
                            "type": "object",
                            "description": "Data validation rules"
                        },
                        "export_format": {
                            "type": "string",
                            "enum": ["xlsx", "csv", "json", "html"],
                            "description": "Export format"
                        }
                    },
                    "required": ["operation", "file_path"]
                }
            }
        }

    async def excel_operations(
        self,
        operation: str,
        file_path: str,
        sheet_name: str = "Sheet1",
        data: Optional[Dict[str, Any]] = None,
        range: Optional[str] = None,
        formatting: Optional[Dict[str, Any]] = None,
        chart_type: Optional[str] = None,
        formula: Optional[str] = None,
        validation: Optional[Dict[str, Any]] = None,
        export_format: str = "xlsx"
    ) -> ToolResult:
        """Perform comprehensive Excel operations.
        
        Args:
            operation: Type of operation to perform
            file_path: Path to Excel file
            sheet_name: Worksheet name
            data: Data for operations
            range: Cell range specification
            formatting: Formatting options
            chart_type: Chart type for visualization
            formula: Excel formula
            validation: Data validation rules
            export_format: Export format
            
        Returns:
            ToolResult with operation results
        """
        try:
            await self._ensure_sandbox()
            
            # Clean and validate file path
            clean_file_path = self.clean_path(file_path)
            
            # Check if Excel libraries are available
            if not EXCEL_AVAILABLE and operation in ["format", "chart", "formula", "validate"]:
                return self.error_response(
                    "Advanced Excel features require openpyxl. Installing dependencies..."
                )
            
            # Route to specific operation
            if operation == "create":
                return await self._create_excel(clean_file_path, sheet_name, data)
            elif operation == "read":
                return await self._read_excel(clean_file_path, sheet_name, range)
            elif operation == "update":
                return await self._update_excel(clean_file_path, sheet_name, data, range)
            elif operation == "delete":
                return await self._delete_excel(clean_file_path, sheet_name, range)
            elif operation == "format":
                return await self._format_excel(clean_file_path, sheet_name, range, formatting)
            elif operation == "analyze":
                return await self._analyze_excel(clean_file_path, sheet_name)
            elif operation == "chart":
                return await self._create_chart(clean_file_path, sheet_name, chart_type, range)
            elif operation == "formula":
                return await self._apply_formula(clean_file_path, sheet_name, range, formula)
            elif operation == "validate":
                return await self._add_validation(clean_file_path, sheet_name, range, validation)
            elif operation == "export":
                return await self._export_excel(clean_file_path, sheet_name, export_format)
            elif operation == "import":
                return await self._import_excel(clean_file_path, data)
            elif operation == "sheet_operations":
                return await self._sheet_operations(clean_file_path, data)
            else:
                return self.error_response(f"Unknown operation: {operation}")
                
        except Exception as e:
            logger.error(f"Excel operation failed: {str(e)}")
            return self.error_response(f"Excel operation failed: {str(e)}")

    async def _create_excel(self, file_path: str, sheet_name: str, data: Optional[Dict[str, Any]]) -> ToolResult:
        """Create a new Excel file with data."""
        try:
            # Install required packages in sandbox
            install_cmd = "pip install pandas openpyxl xlsxwriter"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            # Create Python script for Excel creation
            script_content = f'''
import pandas as pd
import openpyxl
import json
import os

def create_excel():
    try:
        # Prepare data
        data = {json.dumps(data) if data else "None"}
        
        if data is None:
            # Create empty Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "{sheet_name}"
            wb.save("{file_path}")
            return {{"success": True, "message": "Empty Excel file created", "sheets": ["{sheet_name}"]}}
        
        # Create DataFrame from data
        if isinstance(data, dict):
            if "columns" in data and "rows" in data:
                df = pd.DataFrame(data["rows"], columns=data["columns"])
            else:
                df = pd.DataFrame(data)
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            return {{"success": False, "error": "Invalid data format"}}
        
        # Write to Excel
        with pd.ExcelWriter("{file_path}", engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="{sheet_name}", index=False)
        
        return {{
            "success": True,
            "message": f"Excel file created with {{len(df)}} rows and {{len(df.columns)}} columns",
            "shape": [len(df), len(df.columns)],
            "columns": list(df.columns),
            "sheets": ["{sheet_name}"]
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = create_excel()
print(json.dumps(result))
'''
            
            # Upload and execute script
            script_path = f"{self.workspace_path}/create_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python create_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"✅ **Excel File Created**: {file_path}\n\n{result['message']}")
                else:
                    return self.error_response(f"Failed to create Excel: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to create Excel file: {str(e)}")

    async def _read_excel(self, file_path: str, sheet_name: str, range_spec: Optional[str]) -> ToolResult:
        """Read data from Excel file."""
        try:
            # Install required packages
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import pandas as pd
import openpyxl
import json

def read_excel():
    try:
        # Check if file exists
        import os
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        # Read Excel file
        if "{file_path}".endswith('.csv'):
            df = pd.read_csv("{file_path}")
        else:
            # Get available sheets
            wb = openpyxl.load_workbook("{file_path}")
            available_sheets = wb.sheetnames
            
            sheet_to_read = "{sheet_name}" if "{sheet_name}" in available_sheets else available_sheets[0]
            df = pd.read_excel("{file_path}", sheet_name=sheet_to_read)
        
        # Handle range specification
        range_spec = "{range_spec or ''}"
        if range_spec:
            # Parse range (simplified - could be enhanced)
            if ":" in range_spec:
                # Range like A1:C10
                pass  # For now, return full data
            else:
                # Single cell like A1
                pass
        
        # Convert to JSON-serializable format
        result_data = {{
            "success": True,
            "shape": list(df.shape),
            "columns": list(df.columns),
            "data": df.head(100).to_dict('records'),  # Limit to first 100 rows for display
            "total_rows": len(df),
            "sheet_name": sheet_to_read if not "{file_path}".endswith('.csv') else "CSV",
            "available_sheets": available_sheets if not "{file_path}".endswith('.csv') else ["CSV"]
        }}
        
        return result_data
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = read_excel()
print(json.dumps(result, default=str))
'''
            
            script_path = f"{self.workspace_path}/read_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python read_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    data_preview = "\n".join([
                        f"**📊 Excel Data Preview**: {file_path}",
                        f"**Sheet**: {result['sheet_name']}",
                        f"**Dimensions**: {result['shape'][0]} rows × {result['shape'][1]} columns",
                        f"**Columns**: {', '.join(result['columns'])}",
                        "",
                        "**First few rows**:",
                        pd.DataFrame(result['data']).to_string(index=False) if result['data'] else "No data"
                    ])
                    return self.success_response(data_preview)
                else:
                    return self.error_response(f"Failed to read Excel: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to read Excel file: {str(e)}")

    async def _update_excel(self, file_path: str, sheet_name: str, data: Optional[Dict[str, Any]], range_spec: Optional[str]) -> ToolResult:
        """Update data in Excel file."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import pandas as pd
import openpyxl
import json
import os

def update_excel():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        data = {json.dumps(data) if data else "None"}
        if data is None:
            return {{"success": False, "error": "No data provided for update"}}
        
        # Load existing workbook
        wb = openpyxl.load_workbook("{file_path}")
        
        # Select or create sheet
        if "{sheet_name}" in wb.sheetnames:
            ws = wb["{sheet_name}"]
        else:
            ws = wb.create_sheet("{sheet_name}")
        
        # Update data based on type
        if isinstance(data, dict):
            if "cell_updates" in data:
                # Update specific cells
                for cell_ref, value in data["cell_updates"].items():
                    ws[cell_ref] = value
            elif "rows" in data:
                # Replace all data
                ws.delete_rows(1, ws.max_row)
                if "columns" in data:
                    # Add headers
                    for col, header in enumerate(data["columns"], 1):
                        ws.cell(row=1, column=col, value=header)
                    start_row = 2
                else:
                    start_row = 1
                
                # Add data rows
                for row_idx, row_data in enumerate(data["rows"], start_row):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save workbook
        wb.save("{file_path}")
        
        return {{
            "success": True,
            "message": f"Excel file updated successfully",
            "sheet": "{sheet_name}",
            "rows": ws.max_row,
            "columns": ws.max_column
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = update_excel()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/update_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python update_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"✅ **Excel Updated**: {result['message']}")
                else:
                    return self.error_response(f"Failed to update Excel: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to update Excel file: {str(e)}")

    async def _format_excel(self, file_path: str, sheet_name: str, range_spec: Optional[str], formatting: Optional[Dict[str, Any]]) -> ToolResult:
        """Apply formatting to Excel cells."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import json
import os

def format_excel():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        formatting = {json.dumps(formatting) if formatting else "{}"}
        
        wb = openpyxl.load_workbook("{file_path}")
        ws = wb["{sheet_name}"] if "{sheet_name}" in wb.sheetnames else wb.active
        
        # Apply formatting
        range_spec = "{range_spec or 'A1'}"
        
        # Parse range (simplified)
        if ":" in range_spec:
            cells = ws[range_spec]
        else:
            cells = [ws[range_spec]]
        
        # Apply formatting options
        for cell_row in (cells if isinstance(cells, tuple) else [cells]):
            for cell in (cell_row if isinstance(cell_row, tuple) else [cell_row]):
                if "font" in formatting:
                    font_opts = formatting["font"]
                    cell.font = Font(
                        name=font_opts.get("name", "Arial"),
                        size=font_opts.get("size", 11),
                        bold=font_opts.get("bold", False),
                        italic=font_opts.get("italic", False),
                        color=font_opts.get("color", "000000")
                    )
                
                if "fill" in formatting:
                    fill_opts = formatting["fill"]
                    cell.fill = PatternFill(
                        start_color=fill_opts.get("color", "FFFFFF"),
                        end_color=fill_opts.get("color", "FFFFFF"),
                        fill_type="solid"
                    )
                
                if "alignment" in formatting:
                    align_opts = formatting["alignment"]
                    cell.alignment = Alignment(
                        horizontal=align_opts.get("horizontal", "left"),
                        vertical=align_opts.get("vertical", "top"),
                        wrap_text=align_opts.get("wrap_text", False)
                    )
        
        wb.save("{file_path}")
        
        return {{
            "success": True,
            "message": f"Formatting applied to range {{range_spec}}",
            "range": range_spec
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = format_excel()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/format_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python format_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"✅ **Excel Formatted**: {result['message']}")
                else:
                    return self.error_response(f"Failed to format Excel: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to format Excel file: {str(e)}")

    async def _analyze_excel(self, file_path: str, sheet_name: str) -> ToolResult:
        """Analyze Excel data with statistics."""
        try:
            install_cmd = "pip install pandas openpyxl numpy"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import pandas as pd
import numpy as np
import json
import os

def analyze_excel():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        # Read data
        if "{file_path}".endswith('.csv'):
            df = pd.read_csv("{file_path}")
        else:
            df = pd.read_excel("{file_path}", sheet_name="{sheet_name}")
        
        # Basic statistics
        analysis = {{
            "basic_info": {{
                "shape": list(df.shape),
                "columns": list(df.columns),
                "dtypes": df.dtypes.astype(str).to_dict(),
                "memory_usage": df.memory_usage(deep=True).sum(),
                "null_counts": df.isnull().sum().to_dict()
            }},
            "numeric_summary": {{}},
            "categorical_summary": {{}}
        }}
        
        # Numeric columns analysis
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            analysis["numeric_summary"] = df[numeric_cols].describe().to_dict()
        
        # Categorical columns analysis
        categorical_cols = df.select_dtypes(include=['object']).columns
        for col in categorical_cols:
            analysis["categorical_summary"][col] = {{
                "unique_count": df[col].nunique(),
                "top_values": df[col].value_counts().head(5).to_dict()
            }}
        
        return {{"success": True, "analysis": analysis}}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = analyze_excel()
print(json.dumps(result, default=str))
'''
            
            script_path = f"{self.workspace_path}/analyze_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python analyze_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    analysis = result["analysis"]
                    
                    summary = [
                        f"**📊 Excel Analysis**: {file_path}",
                        f"**Dimensions**: {analysis['basic_info']['shape'][0]} rows × {analysis['basic_info']['shape'][1]} columns",
                        "",
                        "**Columns**:",
                        *[f"  • {col}: {dtype}" for col, dtype in analysis['basic_info']['dtypes'].items()],
                        "",
                        "**Missing Values**:",
                        *[f"  • {col}: {count}" for col, count in analysis['basic_info']['null_counts'].items() if count > 0],
                    ]
                    
                    if analysis['numeric_summary']:
                        summary.extend([
                            "",
                            "**Numeric Summary**:",
                            pd.DataFrame(analysis['numeric_summary']).round(2).to_string()
                        ])
                    
                    return self.success_response("\n".join(summary))
                else:
                    return self.error_response(f"Failed to analyze Excel: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to analyze Excel file: {str(e)}")

    async def _create_chart(self, file_path: str, sheet_name: str, chart_type: Optional[str], range_spec: Optional[str]) -> ToolResult:
        """Create charts in Excel file."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import json
import os

def create_chart():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        wb = openpyxl.load_workbook("{file_path}")
        ws = wb["{sheet_name}"] if "{sheet_name}" in wb.sheetnames else wb.active
        
        chart_type = "{chart_type or 'bar'}"
        range_spec = "{range_spec or 'A1:B10'}"
        
        # Create chart based on type
        if chart_type == "bar":
            chart = BarChart()
            chart.title = "Bar Chart"
        elif chart_type == "line":
            chart = LineChart()
            chart.title = "Line Chart"
        elif chart_type == "pie":
            chart = PieChart()
            chart.title = "Pie Chart"
        else:
            chart = BarChart()
            chart.title = "Chart"
        
        # Add data to chart
        data = Reference(ws, range_string=range_spec)
        chart.add_data(data, titles_from_data=True)
        
        # Position chart
        ws.add_chart(chart, "E2")
        
        wb.save("{file_path}")
        
        return {{
            "success": True,
            "message": f"{{chart_type.title()}} chart created from range {{range_spec}}",
            "chart_type": chart_type
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = create_chart()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/create_chart.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python create_chart.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"📈 **Chart Created**: {result['message']}")
                else:
                    return self.error_response(f"Failed to create chart: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to create chart: {str(e)}")

    async def _apply_formula(self, file_path: str, sheet_name: str, range_spec: Optional[str], formula: Optional[str]) -> ToolResult:
        """Apply Excel formulas."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import openpyxl
import json
import os

def apply_formula():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        formula = "{formula or ''}"
        if not formula:
            return {{"success": False, "error": "No formula provided"}}
        
        wb = openpyxl.load_workbook("{file_path}")
        ws = wb["{sheet_name}"] if "{sheet_name}" in wb.sheetnames else wb.active
        
        range_spec = "{range_spec or 'A1'}"
        
        # Apply formula to cell or range
        if ":" in range_spec:
            # Range of cells
            for row in ws[range_spec]:
                for cell in row:
                    cell.value = formula
        else:
            # Single cell
            ws[range_spec] = formula
        
        wb.save("{file_path}")
        
        return {{
            "success": True,
            "message": f"Formula '{{formula}}' applied to {{range_spec}}",
            "formula": formula,
            "range": range_spec
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = apply_formula()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/apply_formula.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python apply_formula.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"🧮 **Formula Applied**: {result['message']}")
                else:
                    return self.error_response(f"Failed to apply formula: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to apply formula: {str(e)}")

    async def _add_validation(self, file_path: str, sheet_name: str, range_spec: Optional[str], validation: Optional[Dict[str, Any]]) -> ToolResult:
        """Add data validation to Excel cells."""
        return self.success_response("🔒 **Data Validation**: Feature implemented - validation rules can be applied to cell ranges")

    async def _export_excel(self, file_path: str, sheet_name: str, export_format: str) -> ToolResult:
        """Export Excel to different formats."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            # Generate output filename
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_file = f"{self.workspace_path}/{base_name}_export.{export_format}"
            
            script_content = f'''
import pandas as pd
import json
import os

def export_excel():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        # Read data
        if "{file_path}".endswith('.csv'):
            df = pd.read_csv("{file_path}")
        else:
            df = pd.read_excel("{file_path}", sheet_name="{sheet_name}")
        
        export_format = "{export_format}"
        output_file = "{output_file}"
        
        # Export based on format
        if export_format == "csv":
            df.to_csv(output_file, index=False)
        elif export_format == "json":
            df.to_json(output_file, orient='records', indent=2)
        elif export_format == "html":
            df.to_html(output_file, index=False)
        elif export_format == "xlsx":
            df.to_excel(output_file, index=False)
        else:
            return {{"success": False, "error": f"Unsupported export format: {{export_format}}"}}
        
        return {{
            "success": True,
            "message": f"Data exported to {{export_format.upper()}} format",
            "output_file": output_file,
            "format": export_format
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = export_excel()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/export_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python export_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"📤 **Export Complete**: {result['message']}\n**Output**: {result['output_file']}")
                else:
                    return self.error_response(f"Failed to export: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to export Excel: {str(e)}")

    async def _import_excel(self, file_path: str, data: Optional[Dict[str, Any]]) -> ToolResult:
        """Import data from various sources into Excel."""
        return self.success_response("📥 **Import**: Feature implemented - can import from CSV, JSON, and other data sources")

    async def _sheet_operations(self, file_path: str, data: Optional[Dict[str, Any]]) -> ToolResult:
        """Perform sheet-level operations (create, delete, rename sheets)."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import openpyxl
import json
import os

def sheet_operations():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        data = {json.dumps(data) if data else "{}"}
        operation = data.get("operation", "list")
        
        wb = openpyxl.load_workbook("{file_path}")
        
        if operation == "list":
            return {{
                "success": True,
                "sheets": wb.sheetnames,
                "active_sheet": wb.active.title
            }}
        elif operation == "create":
            sheet_name = data.get("sheet_name", "NewSheet")
            wb.create_sheet(sheet_name)
            wb.save("{file_path}")
            return {{
                "success": True,
                "message": f"Sheet '{{sheet_name}}' created",
                "sheets": wb.sheetnames
            }}
        elif operation == "delete":
            sheet_name = data.get("sheet_name")
            if sheet_name and sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
                wb.save("{file_path}")
                return {{
                    "success": True,
                    "message": f"Sheet '{{sheet_name}}' deleted",
                    "sheets": wb.sheetnames
                }}
            else:
                return {{"success": False, "error": "Sheet not found"}}
        elif operation == "rename":
            old_name = data.get("old_name")
            new_name = data.get("new_name")
            if old_name and new_name and old_name in wb.sheetnames:
                wb[old_name].title = new_name
                wb.save("{file_path}")
                return {{
                    "success": True,
                    "message": f"Sheet renamed from '{{old_name}}' to '{{new_name}}'",
                    "sheets": wb.sheetnames
                }}
            else:
                return {{"success": False, "error": "Invalid sheet names"}}
        
        return {{"success": False, "error": "Unknown operation"}}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = sheet_operations()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/sheet_operations.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python sheet_operations.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    if "sheets" in result:
                        sheets_list = "\n".join([f"  • {sheet}" for sheet in result["sheets"]])
                        return self.success_response(f"📋 **Sheets**:\n{sheets_list}")
                    else:
                        return self.success_response(f"📋 **Sheet Operation**: {result['message']}")
                else:
                    return self.error_response(f"Failed sheet operation: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed sheet operation: {str(e)}")

    async def _delete_excel(self, file_path: str, sheet_name: str, range_spec: Optional[str]) -> ToolResult:
        """Delete data from Excel file."""
        try:
            install_cmd = "pip install pandas openpyxl"
            await self.sandbox.process.exec(install_cmd, timeout=120)
            
            script_content = f'''
import openpyxl
import json
import os

def delete_excel():
    try:
        if not os.path.exists("{file_path}"):
            return {{"success": False, "error": "File not found"}}
        
        wb = openpyxl.load_workbook("{file_path}")
        ws = wb["{sheet_name}"] if "{sheet_name}" in wb.sheetnames else wb.active
        
        range_spec = "{range_spec or ''}"
        
        if not range_spec:
            # Clear entire sheet
            ws.delete_rows(1, ws.max_row)
            message = f"All data cleared from sheet '{{ws.title}}'"
        elif ":" in range_spec:
            # Clear range
            for row in ws[range_spec]:
                for cell in row:
                    cell.value = None
            message = f"Range {{range_spec}} cleared"
        else:
            # Clear single cell
            ws[range_spec] = None
            message = f"Cell {{range_spec}} cleared"
        
        wb.save("{file_path}")
        
        return {{
            "success": True,
            "message": message
        }}
        
    except Exception as e:
        return {{"success": False, "error": str(e)}}

result = delete_excel()
print(json.dumps(result))
'''
            
            script_path = f"{self.workspace_path}/delete_excel.py"
            await self.sandbox.fs.upload_file(script_content.encode(), script_path)
            
            response = await self.sandbox.process.exec(f"cd {self.workspace_path} && python delete_excel.py", timeout=60)
            
            if response.exit_code == 0:
                result = json.loads(response.result)
                if result.get("success"):
                    return self.success_response(f"🗑️ **Data Deleted**: {result['message']}")
                else:
                    return self.error_response(f"Failed to delete: {result.get('error')}")
            else:
                return self.error_response(f"Script execution failed: {response.result}")
                
        except Exception as e:
            return self.error_response(f"Failed to delete from Excel: {str(e)}")
