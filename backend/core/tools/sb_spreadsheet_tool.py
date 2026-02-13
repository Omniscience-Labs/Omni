"""XLSX tool: JSON workbook model → .xlsx (in-memory only). Full regen per call; only .xlsx is written."""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import jsonschema
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import column_index_from_string

from core.agentpress.thread_manager import ThreadManager
from core.agentpress.tool import ToolResult, openapi_schema, tool_metadata
from core.sandbox.tool_base import SandboxToolsBase
from core.utils.logger import logger
<<<<<<< Updated upstream
from typing import List, Dict, Optional, Union, Any
import json
import os
from datetime import datetime
import re
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import Cell
from pathlib import Path
=======

# ---------------------------------------------------------------------------
# JSON Schema – Foundational v1
# ---------------------------------------------------------------------------

XLSX_WORKBOOK_SCHEMA: Dict[str, Any] = {
    "$schema": "https://json-schema.org/draft/2020-12/schema",
    "$id": "https://example.com/schemas/min-ai-xlsx.schema.json",
    "title": "Minimal AI XLSX Workbook Schema (Foundational v1)",
    "description": (
        "Foundational XLSX schema for JSON-template to .xlsx rendering. "
        "Cell refs are A1-style. Colors are 6- or 8-hex RGB (no #)."
    ),
    "type": "object",
    "required": ["workbook"],
    "additionalProperties": True,
    "properties": {
        "version": {"type": "string", "default": "1.0"},
        "workbook": {
            "type": "object",
            "required": ["sheets"],
            "additionalProperties": True,
            "properties": {
                "title": {"type": "string"},
                "creator": {"type": "string"},
                "activeSheet": {
                    "type": "string",
                    "description": "Name of the sheet to make active (optional).",
                },
                "styles": {
                    "type": "object",
                    "description": "Named style dictionary. Cells reference these by name via 'style'.",
                    "additionalProperties": {"$ref": "#/$defs/CellFormat"},
                    "default": {},
                },
                "sheets": {
                    "type": "array",
                    "minItems": 1,
                    "items": {"$ref": "#/$defs/Sheet"},
                },
            },
        },
    },
    "$defs": {
        "A1CellRef": {
            "type": "string",
            "pattern": "^[A-Z]+[1-9][0-9]*$",
            "description": "Excel A1-style cell reference (e.g. A1, B2, AA10).",
        },
        "A1RangeRef": {
            "type": "string",
            "pattern": "^[A-Z]+[1-9][0-9]*:[A-Z]+[1-9][0-9]*$",
            "description": "Excel A1-style range reference (e.g. A1:B2).",
        },
        "ColorHex": {
            "type": "string",
            "pattern": "^[0-9A-Fa-f]{6}([0-9A-Fa-f]{2})?$",
            "description": "RGB (RRGGBB) or ARGB (AARRGGBB) hex string without #.",
        },
        "CellFormat": {
            "type": "object",
            "additionalProperties": True,
            "properties": {
                "number_format": {
                    "type": "string",
                    "description": (
                        "Excel number format code. Supports dates (mm/dd/yyyy), decimals (0.00), "
                        "currency ($#,##0.00), percentage (0.00%), and unit formatting: use quotes "
                        "for literal text, e.g. 0.00 \"kg\", #,##0 \"m\", 0.0 \"hrs\". See Excel number format syntax."
                    ),
                },
                "bold": {"type": "boolean"},
                "italic": {"type": "boolean"},
                "font_size": {"type": "number", "minimum": 1},
                "font_color": {"$ref": "#/$defs/ColorHex"},
                "fill_color": {"$ref": "#/$defs/ColorHex"},
                "alignment": {
                    "type": "string",
                    "enum": ["left", "center", "right", "top", "bottom", "middle"],
                    "description": "Text alignment.",
                },
                "wrap_text": {"type": "boolean"},
                "border": {
                    "type": "boolean",
                    "description": "If true, apply a thin border around the cell.",
                },
            },
        },
        "Cell": {
            "type": "object",
            "required": ["ref"],
            "additionalProperties": True,
            "properties": {
                "ref": {"$ref": "#/$defs/A1CellRef"},
                "value": {
                    "description": (
                        "Cell value. If value_type is omitted and value is an ISO "
                        "date/datetime string, renderer may parse into an Excel date."
                    ),
                    "oneOf": [
                        {"type": "string"},
                        {"type": "number"},
                        {"type": "boolean"},
                        {"type": "null"},
                    ],
                },
                "formula": {
                    "type": "string",
                    "pattern": "^=",
                    "description": "Excel formula starting with '='. Takes precedence over value.",
                },
                "value_type": {
                    "type": "string",
                    "enum": ["auto", "string", "number", "boolean", "date", "datetime"],
                    "default": "auto",
                    "description": (
                        "Controls parsing/coercion. Use 'string' to prevent '=...' "
                        "being treated as formula when using value only."
                    ),
                },
                "style": {
                    "description": "Named reference into workbook.styles, or an inline format object.",
                    "oneOf": [
                        {"type": "string"},
                        {"$ref": "#/$defs/CellFormat"},
                    ],
                },
            },
        },
        "Sheet": {
            "type": "object",
            "required": ["name"],
            "additionalProperties": True,
            "properties": {
                "name": {"type": "string", "minLength": 1, "maxLength": 31},
                "cells": {
                    "type": "array",
                    "default": [],
                    "items": {"$ref": "#/$defs/Cell"},
                },
                "merges": {
                    "type": "array",
                    "default": [],
                    "items": {"$ref": "#/$defs/A1RangeRef"},
                    "description": "Ranges to merge, e.g. ['A1:D1', 'B2:C2'].",
                },
            },
        },
    },
}

# ---------------------------------------------------------------------------
# A1-reference regex (used by semantic validation)
# ---------------------------------------------------------------------------

_A1_CELL_RE = re.compile(r"^[A-Z]+[1-9][0-9]*$")
_A1_RANGE_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*):([A-Z]+)([1-9][0-9]*)$")


# ---------------------------------------------------------------------------
# Normalization (mirrors _normalize_document_model in sb_docx_tool)
# ---------------------------------------------------------------------------


def _normalize_workbook_model(workbook_json: Dict[str, Any]) -> Dict[str, Any]:
    """
    Accept several input shapes and normalize to the canonical schema form:
      - canonical: {"workbook": {"sheets": [...], "styles": {...}}}
      - legacy:    {"sheets": [...], "styles": {...}}          → wraps into workbook
      - alias:     cell-level "format" key                     → mapped to "style"
    Never mutates the caller's data (shallow copy).
    """
    if not isinstance(workbook_json, dict):
        raise ValueError("workbook_json must be an object")

    normalized: Dict[str, Any] = dict(workbook_json)

    # ------------------------------------------------------------------
    # Wrap top-level sheets/styles into workbook if missing
    # ------------------------------------------------------------------
    wb_obj = normalized.get("workbook")
    if not isinstance(wb_obj, dict):
        wb_obj = {}

    # top-level "sheets" → workbook.sheets
    if "sheets" not in wb_obj and isinstance(normalized.get("sheets"), list):
        wb_obj["sheets"] = normalized.pop("sheets")

    # top-level "styles" → workbook.styles
    if "styles" not in wb_obj and isinstance(normalized.get("styles"), dict):
        wb_obj["styles"] = normalized.pop("styles")

    # top-level metadata shortcuts
    for key in ("title", "creator", "activeSheet"):
        if key not in wb_obj and key in normalized:
            wb_obj[key] = normalized.pop(key)

    normalized["workbook"] = wb_obj

    # ------------------------------------------------------------------
    # Cell-level alias: "format" → "style"
    # ------------------------------------------------------------------
    sheets = wb_obj.get("sheets")
    if isinstance(sheets, list):
        for sheet in sheets:
            if not isinstance(sheet, dict):
                continue
            cells = sheet.get("cells")
            if not isinstance(cells, list):
                continue
            for cell in cells:
                if not isinstance(cell, dict):
                    continue
                if "style" not in cell and "format" in cell:
                    cell["style"] = cell.pop("format")

    return normalized


# ---------------------------------------------------------------------------
# Semantic validation (beyond JSON Schema)
# ---------------------------------------------------------------------------


def _validate_workbook_semantics(normalized: Dict[str, Any]) -> Optional[str]:
    """
    Return an error message string if something is semantically wrong,
    or None if everything looks good.
    """
    wb = normalized.get("workbook") or {}
    sheets = wb.get("sheets") or []
    styles = wb.get("styles") or {}

    if not sheets:
        return "Workbook must contain at least one sheet."

    # --- Sheet-name checks ---
    seen_names: set[str] = set()
    for idx, sheet in enumerate(sheets):
        name = sheet.get("name")
        if not name or not isinstance(name, str):
            return f"Sheet at index {idx} must have a non-empty 'name'."
        if len(name) > 31:
            return f"Sheet name '{name[:34]}…' exceeds Excel's 31-character limit."
        if name in seen_names:
            return f"Duplicate sheet name '{name}' at index {idx}."
        seen_names.add(name)

    # --- Per-sheet cell / merge checks ---
    for sheet in sheets:
        sheet_name = sheet.get("name", "?")

        # Cells
        for ci, cell in enumerate(sheet.get("cells") or []):
            ref = cell.get("ref", "")
            if not _A1_CELL_RE.match(ref):
                return f"Sheet '{sheet_name}', cell index {ci}: invalid A1 ref '{ref}'."

            # Formula must start with '='
            formula = cell.get("formula")
            if formula is not None and not str(formula).startswith("="):
                return (
                    f"Sheet '{sheet_name}', cell {ref}: formula must start with '='. "
                    f"Got: '{str(formula)[:40]}'."
                )

            # Style reference must exist in named styles
            style_val = cell.get("style")
            if isinstance(style_val, str) and style_val not in styles:
                return (
                    f"Sheet '{sheet_name}', cell {ref}: style '{style_val}' not found "
                    f"in workbook.styles. Available: {list(styles.keys())}."
                )

        # Merges
        for mi, merge in enumerate(sheet.get("merges") or []):
            m = _A1_RANGE_RE.match(merge)
            if not m:
                return f"Sheet '{sheet_name}', merge index {mi}: invalid range '{merge}'."
            # Verify end >= start (column and row)
            start_col, start_row, end_col, end_row = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            try:
                if column_index_from_string(end_col) < column_index_from_string(start_col):
                    return (
                        f"Sheet '{sheet_name}', merge '{merge}': end column ({end_col}) "
                        f"is before start column ({start_col})."
                    )
            except ValueError:
                return f"Sheet '{sheet_name}', merge '{merge}': invalid column letter."
            if end_row < start_row:
                return (
                    f"Sheet '{sheet_name}', merge '{merge}': end row ({end_row}) "
                    f"is before start row ({start_row})."
                )

    return None


# ---------------------------------------------------------------------------
# Path / filename helpers  (mirrors sb_docx_tool)
# ---------------------------------------------------------------------------


def _validate_and_normalize_path(workspace_path: str, path: str) -> Tuple[str, str]:
    if not path or not isinstance(path, str):
        raise ValueError("Path must be a non-empty string")
    path = path.replace("\\", "/").strip()
    if path.startswith("/workspace/"):
        path = path[len("/workspace/"):].lstrip("/")
    parts = path.split("/")
    normalized: List[str] = []
    for part in parts:
        if part == "..":
            if normalized:
                normalized.pop()
        elif part and part != ".":
            normalized.append(part)
    cleaned = "/".join(normalized)
    if ".." in cleaned or "\\" in cleaned:
        raise ValueError("Path cannot contain '..' or backslashes")
    return cleaned, f"{workspace_path}/{cleaned}"


def _sanitize_filename(name: str) -> str:
    base = "workbook"
    if name:
        safe = "".join(c for c in name if c.isalnum() or c in "-_.").strip()
        if safe and not safe.startswith("."):
            base = safe
    if not base.lower().endswith(".xlsx"):
        base = base.rstrip(".") + ".xlsx"
    return base


# ---------------------------------------------------------------------------
# Cell value parsing
# ---------------------------------------------------------------------------

_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_ISO_DATETIME_RE = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}(:\d{2})?")


def _parse_cell_value(value: Any, value_type: str = "auto") -> Any:
    """
    Parse a raw JSON value into the appropriate Python type for openpyxl.

    value_type controls coercion:
      - "auto"     : attempt smart type detection (default)
      - "string"   : always store as str (prevents formula injection)
      - "number"   : coerce to int/float
      - "boolean"  : coerce to bool
      - "date"     : parse ISO date string → datetime.date
      - "datetime" : parse ISO datetime string → datetime.datetime
    """
    if value is None:
        return None

    vt = str(value_type).lower() if value_type else "auto"

    if vt == "string":
        return str(value)

    if vt == "number":
        try:
            if isinstance(value, (int, float)):
                return value
            s = str(value)
            return float(s) if "." in s else int(s)
        except (TypeError, ValueError):
            return str(value)

    if vt == "boolean":
        if isinstance(value, bool):
            return value
        return bool(value)

    if vt in ("date", "datetime"):
        if isinstance(value, str):
            try:
                if "T" in value or " " in value:
                    return datetime.fromisoformat(value.replace("Z", "+00:00"))
                return datetime.strptime(value, "%Y-%m-%d")
            except (ValueError, TypeError):
                pass
        return value

    # "auto" — smart detection
    if isinstance(value, (int, float, bool)):
        return value

    if isinstance(value, str):
        # ISO date/datetime
        if _ISO_DATETIME_RE.match(value):
            try:
                return datetime.fromisoformat(value.replace("Z", "+00:00"))
            except (ValueError, TypeError):
                pass
        elif _ISO_DATE_RE.match(value):
            try:
                return datetime.strptime(value, "%Y-%m-%d")
            except (ValueError, TypeError):
                pass
        # Numeric strings
        try:
            return float(value) if "." in value else int(value)
        except ValueError:
            pass
        return value

    return str(value)


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

_ALIGNMENT_MAP = {
    "left": Alignment(horizontal="left"),
    "center": Alignment(horizontal="center"),
    "right": Alignment(horizontal="right"),
    "top": Alignment(vertical="top"),
    "bottom": Alignment(vertical="bottom"),
    "middle": Alignment(vertical="center"),
}

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _resolve_style(
    cell_style: Any,
    named_styles: Dict[str, Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    """Resolve a cell's style field to a concrete format dict (or None)."""
    if isinstance(cell_style, dict):
        return cell_style
    if isinstance(cell_style, str):
        return named_styles.get(cell_style)
    return None


def _apply_cell_format(ws_cell: Any, fmt: Dict[str, Any]) -> None:
    """Apply a CellFormat dict to an openpyxl cell."""
    if not fmt or not isinstance(fmt, dict):
        return

    # Number format
    nf = fmt.get("number_format")
    if nf is not None:
        ws_cell.number_format = str(nf)

    # Font properties
    font_kwargs: Dict[str, Any] = {}
    if fmt.get("bold") is not None:
        font_kwargs["bold"] = bool(fmt["bold"])
    if fmt.get("italic") is not None:
        font_kwargs["italic"] = bool(fmt["italic"])
    if fmt.get("font_size") is not None:
        font_kwargs["size"] = float(fmt["font_size"])
    fc = fmt.get("font_color")
    if fc:
        font_kwargs["color"] = Color(rgb=str(fc))
    if font_kwargs:
        # Merge with existing font to preserve defaults
        existing = ws_cell.font
        ws_cell.font = Font(
            bold=font_kwargs.get("bold", existing.bold),
            italic=font_kwargs.get("italic", existing.italic),
            size=font_kwargs.get("size", existing.size),
            color=font_kwargs.get("color", existing.color),
            name=existing.name,
        )

    # Fill color
    fill_c = fmt.get("fill_color")
    if fill_c:
        ws_cell.fill = PatternFill(start_color=str(fill_c), end_color=str(fill_c), fill_type="solid")

    # Alignment
    align_key = fmt.get("alignment")
    wrap = fmt.get("wrap_text")
    if align_key or wrap is not None:
        base = _ALIGNMENT_MAP.get(str(align_key).lower()) if align_key else Alignment()
        # Combine alignment + wrap_text
        ws_cell.alignment = Alignment(
            horizontal=base.horizontal,
            vertical=base.vertical,
            wrap_text=bool(wrap) if wrap is not None else base.wrap_text,
        )

    # Border
    if fmt.get("border"):
        ws_cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Workbook renderer
# ---------------------------------------------------------------------------


def _render_workbook(normalized: Dict[str, Any]) -> Tuple[Workbook, int, int, int]:
    """
    Render a normalized+validated workbook model into an openpyxl Workbook.

    Returns:
        (workbook, sheet_count, cell_count, merge_count)
    """
    wb_model = normalized.get("workbook") or {}
    sheets_model = wb_model.get("sheets") or []
    named_styles = wb_model.get("styles") or {}

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    total_cells = 0
    total_merges = 0

    for sheet_model in sheets_model:
        ws = wb.create_sheet(title=sheet_model["name"])

        # --- Merges (apply first so merged area exists before writing) ---
        merges = sheet_model.get("merges") or []
        for merge_range in merges:
            ws.merge_cells(merge_range)
        total_merges += len(merges)

        # --- Cells ---
        cells = sheet_model.get("cells") or []
        for cell_model in cells:
            ref = cell_model["ref"]
            ws_cell = ws[ref]

            # Value / formula
            formula = cell_model.get("formula")
            if formula is not None:
                ws_cell.value = str(formula)
            else:
                raw_value = cell_model.get("value")
                vtype = cell_model.get("value_type", "auto")
                ws_cell.value = _parse_cell_value(raw_value, vtype)

            # Formatting
            fmt = _resolve_style(cell_model.get("style"), named_styles)
            if fmt:
                _apply_cell_format(ws_cell, fmt)

            total_cells += 1

    # --- Workbook metadata ---
    title = wb_model.get("title")
    if title:
        wb.properties.title = str(title)
    creator = wb_model.get("creator")
    if creator:
        wb.properties.creator = str(creator)

    # --- Active sheet ---
    active_name = wb_model.get("activeSheet")
    if active_name and active_name in wb.sheetnames:
        wb.active = wb.sheetnames.index(active_name)

    return wb, len(sheets_model), total_cells, total_merges


# ---------------------------------------------------------------------------
# Tool class
# ---------------------------------------------------------------------------
>>>>>>> Stashed changes


@tool_metadata(
    display_name="Spreadsheets",
    description="Generate Excel (.xlsx) workbooks from a JSON workbook model (single render per request)",
    icon="FileSpreadsheet",
    color="bg-green-100 dark:bg-green-800/50",
    weight=80,
    visible=True,
)
class SandboxSpreadsheetTool(SandboxToolsBase):
    def __init__(self, project_id: str, thread_manager: ThreadManager):
        super().__init__(project_id, thread_manager)
        self.spreadsheets_dir = "spreadsheets"

    async def _ensure_spreadsheets_dir(self) -> None:
        try:
            await self.sandbox.fs.create_folder(f"{self.workspace_path}/{self.spreadsheets_dir}", "755")
        except Exception:
            pass

    @openapi_schema({
        "type": "function",
        "function": {
            "name": "render_xlsx",
            "description": (
                "Render a workbook from a JSON workbook model and save as .xlsx in the workspace. "
                "The JSON is not stored; only the generated .xlsx file is written. "
                "Full regen per call — the entire workbook is rebuilt from the template each time."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "workbook_json": {
                        "type": "object",
                        "description": (
                            "The workbook model. Must contain 'workbook.sheets' (array of sheet objects). "
                            "Each sheet has 'name', optional 'cells' array (each cell has 'ref', optional "
                            "'value'/'formula'/'style'), and optional 'merges' array of A1 range strings. "
                            "Optional 'workbook.styles' for named style definitions."
                        ),
                    },
                    "output_filename": {
                        "type": "string",
                        "description": "Output filename (will be forced to .xlsx extension).",
                        "default": "workbook.xlsx",
                    },
                },
                "required": ["workbook_json"],
            },
        },
    })
    async def render_xlsx(
        self,
        workbook_json: Dict[str, Any],
        output_filename: str = "workbook.xlsx",
    ) -> ToolResult:
        try:
            await self._ensure_sandbox()
            await self._ensure_spreadsheets_dir()

            # --- Normalize ---
            try:
                normalized = _normalize_workbook_model(workbook_json)
            except ValueError as e:
                return self.fail_response(str(e))

            # --- JSON Schema validation ---
            try:
                jsonschema.validate(instance=normalized, schema=XLSX_WORKBOOK_SCHEMA)
            except jsonschema.ValidationError as e:
                return self.fail_response(f"Workbook model validation failed: {e.message or str(e)}")

            # --- Semantic validation ---
            semantic_err = _validate_workbook_semantics(normalized)
            if semantic_err:
                return self.fail_response(f"Workbook model invalid: {semantic_err}")

            # --- Render ---
            wb, sheet_count, cell_count, merge_count = _render_workbook(normalized)

            # --- Write to sandbox ---
            safe_name = _sanitize_filename(output_filename)
            relative_path = f"{self.spreadsheets_dir}/{safe_name}"
            cleaned, full_path = _validate_and_normalize_path(self.workspace_path, relative_path)

            # Delete existing file if present (full regen)
            try:
                await self.sandbox.fs.delete_file(full_path)
            except Exception:
                pass

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            await self.sandbox.fs.upload_file(buf.getvalue(), full_path)

            logger.info(
                "sb_spreadsheet_tool render_xlsx: sheets=%s cells=%s merges=%s path=%s",
                sheet_count, cell_count, merge_count, full_path,
            )
            return self.success_response({
                "saved_path": full_path,
                "relative_path": f"/workspace/{cleaned}",
                "sheet_count": sheet_count,
                "cell_count": cell_count,
                "merge_count": merge_count,
            })

        except ValueError as e:
            return self.fail_response(str(e))
        except Exception as e:
<<<<<<< Updated upstream
            logger.error(f"Error adding sheet: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to add sheet: {str(e)}")
    
    # ==================== FUNCTION 3: delete_sheet ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "delete_sheet",
            "description": "Delete a worksheet from a workbook. If deleting the last sheet, automatically creates a default 'Sheet1' to satisfy Excel's requirement of at least one sheet.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet to delete"
                    }
                },
                "required": ["path", "sheet_name"]
            }
        }
    })
    async def delete_sheet(
        self,
        path: str,
        sheet_name: str
    ) -> ToolResult:
        """Delete a worksheet from workbook"""
        try:
            await self._ensure_sandbox()
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Check if sheet exists
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            
            # Protect against deleting last sheet
            if len(workbook.sheetnames) == 1:
                # Auto-create default sheet before deleting
                workbook.create_sheet(title="Sheet1")
                logger.info(f"Auto-created default 'Sheet1' before deleting last sheet")
            
            # Delete sheet
            del workbook[sheet_name]
            
            # Save back
            await self._save_workbook_to_sandbox(workbook, full_path)
            
            logger.info(f"Deleted sheet '{sheet_name}' from {full_path}")
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "sheet_names": workbook.sheetnames
            })
            
        except Exception as e:
            logger.error(f"Error deleting sheet: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to delete sheet: {str(e)}")
    
    # ==================== FUNCTION 4: update_cell ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "update_cell",
            "description": "Update a single cell in a spreadsheet with a value. Supports string, number, bool, null, and ISO date strings.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "cell": {
                        "type": "string",
                        "description": "Cell reference in A1 format (e.g., 'A1', 'B5')"
                    },
                    "value": {
                        "type": ["string", "number", "boolean", "null"],
                        "description": "Value to set. Can be string, number, bool, null, or ISO date string (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)"
                    }
                },
                "required": ["path", "sheet_name", "cell", "value"]
            }
        }
    })
    async def update_cell(
        self,
        path: str,
        sheet_name: str,
        cell: str,
        value: Union[str, int, float, bool, None]
    ) -> ToolResult:
        """Update a single cell"""
        try:
            await self._ensure_sandbox()
            
            # Validate cell reference
            if not self._validate_cell_reference(cell):
                return self.fail_response(f"Invalid cell reference: {cell}. Expected format: A1")
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            
            # Update cell
            cell_obj = worksheet[cell.upper()]
            cell_obj.value = self._parse_cell_value(value)
            
            # Save back
            await self._save_workbook_to_sandbox(workbook, full_path)
            
            logger.info(f"Updated cell {cell} in {full_path}")
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "updated_cell": cell.upper(),
                "value": str(value) if value is not None else None
            })
            
        except Exception as e:
            logger.error(f"Error updating cell: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to update cell: {str(e)}")
    
    # ==================== FUNCTION 5: update_range ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "update_range",
            "description": "Update a range of cells with values from a 2D array. Use 'range' parameter (A1:C3 format) to specify the target range.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "range": {
                        "type": "string",
                        "description": "Cell range in A1:C3 format. The top-left cell of this range is where writing starts; data size can be any number of rows/columns."
                    },
                    "values": {
                        "type": "array",
                        "description": "2D array of values (rows x cols) to write. Writing starts at the top-left of the range and extends to fit all provided rows and columns.",
                        "items": {
                            "type": "array",
                            "items": {
                                "type": ["string", "number", "boolean", "null"]
                            }
                        }
                    }
                },
                "required": ["path", "sheet_name", "range", "values"]
            }
        }
    })
    async def update_range(
        self,
        path: str,
        sheet_name: str,
        range: str,
        values: List[List[Any]]
    ) -> ToolResult:
        """Update a range of cells"""
        try:
            await self._ensure_sandbox()
            
            # Validate range
            if not self._validate_range(range):
                return self.fail_response(f"Invalid range format: {range}. Expected format: A1:C3")
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            
            # Parse range (used as top-left anchor; data extent can be larger or smaller)
            start_row, start_col, end_row, end_col = self._parse_range(range)
            
            actual_rows = len(values)
            actual_cols = max(len(row) for row in values) if values else 0
            
            if actual_rows == 0:
                return self.fail_response("No data to write (values array is empty)")
            
            # Write data
            for row_idx, row_data in enumerate(values):
                for col_idx, value in enumerate(row_data):
                    cell = worksheet.cell(row=start_row + row_idx, column=start_col + col_idx)
                    cell.value = self._parse_cell_value(value)
            
            # Save back
            await self._save_workbook_to_sandbox(workbook, full_path)
            
            end_row_written = start_row + actual_rows - 1
            end_col_written = start_col + actual_cols - 1
            range_written = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col_written)}{end_row_written}"
            logger.info(f"Updated range {range_written} in {full_path}")
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "range_written": range_written,
                "rows": actual_rows,
                "cols": actual_cols
            })
            
        except Exception as e:
            logger.error(f"Error updating range: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to update range: {str(e)}")
    
    # ==================== clear_range ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "clear_range",
            "description": "Clear cell values in a range (set to empty). Use after read_range + update_range to complete a move, or to clear contents without deleting the sheet.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to existing workbook file"},
                    "sheet_name": {"type": "string", "description": "Name of the sheet"},
                    "range": {"type": "string", "description": "Range to clear in A1:C3 format (e.g. A1:B10)"}
                },
                "required": ["path", "sheet_name", "range"]
            }
        }
    })
    async def clear_range(
        self,
        path: str,
        sheet_name: str,
        range: str
    ) -> ToolResult:
        """Clear cell values in a range."""
        try:
            await self._ensure_sandbox()
            if not self._validate_range(range):
                return self.fail_response(f"Invalid range format: {range}. Expected format: A1:C3")
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            workbook = await self._load_workbook_from_sandbox(full_path)
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            start_row, start_col, end_row, end_col = self._parse_range(range)
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    worksheet.cell(row=r, column=c).value = None
            await self._save_workbook_to_sandbox(workbook, full_path)
            rows = end_row - start_row + 1
            cols = end_col - start_col + 1
            logger.info(f"Cleared range {range} in {full_path}")
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "range_cleared": range,
                "rows": rows,
                "cols": cols
            })
        except Exception as e:
            logger.error(f"Error clearing range: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to clear range: {str(e)}")
    
    # ==================== FUNCTION 6: read_cell ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "read_cell",
            "description": "Read a single cell value from a spreadsheet. Returns the value and formula if present.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "cell": {
                        "type": "string",
                        "description": "Cell reference in A1 format (e.g., 'A1', 'B5')"
                    }
                },
                "required": ["path", "sheet_name", "cell"]
            }
        }
    })
    async def read_cell(
        self,
        path: str,
        sheet_name: str,
        cell: str
    ) -> ToolResult:
        """Read a single cell value"""
        try:
            await self._ensure_sandbox()
            
            # Validate cell reference
            if not self._validate_cell_reference(cell):
                return self.fail_response(f"Invalid cell reference: {cell}. Expected format: A1")
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            
            # Read cell
            cell_obj = worksheet[cell.upper()]
            
            # Get formula if present (openpyxl stores formulas separately)
            formula = None
            value = cell_obj.value
            
            # Check if cell contains a formula
            if hasattr(cell_obj, 'data_type') and cell_obj.data_type == 'f':
                # Cell contains a formula - value is the formula string
                formula = str(cell_obj.value) if cell_obj.value else None
                # For formulas, we return the formula string as the value
                # (openpyxl doesn't compute formula results without Excel engine)
                value = formula
            else:
                # Regular cell value - format for output
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif value is None:
                    value = None
                else:
                    value = str(value)
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "value": value,
                "formula": formula
            })
            
        except Exception as e:
            logger.error(f"Error reading cell: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to read cell: {str(e)}")
    
    # ==================== FUNCTION 7: read_range ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "read_range",
            "description": "Read a range of cells from a spreadsheet. Returns a 2D array of values.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "range": {
                        "type": "string",
                        "description": "Cell range in A1:C10 format"
                    }
                },
                "required": ["path", "sheet_name", "range"]
            }
        }
    })
    async def read_range(
        self,
        path: str,
        sheet_name: str,
        range: str
    ) -> ToolResult:
        """Read a range of cells"""
        try:
            await self._ensure_sandbox()
            
            # Validate range
            if not self._validate_range(range):
                return self.fail_response(f"Invalid range format: {range}. Expected format: A1:C10")
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            
            # Parse range
            start_row, start_col, end_row, end_col = self._parse_range(range)
            
            # Read data
            values = []
            for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True):
                row_values = []
                for cell_value in row:
                    if isinstance(cell_value, datetime):
                        row_values.append(cell_value.isoformat())
                    elif cell_value is None:
                        row_values.append(None)
                    else:
                        row_values.append(str(cell_value))
                values.append(row_values)
            
            rows = len(values)
            cols = len(values[0]) if values else 0
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "values": values,
                "rows": rows,
                "cols": cols
            })
            
        except Exception as e:
            logger.error(f"Error reading range: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to read range: {str(e)}")
    
    # ==================== FUNCTION 8: read_spreadsheet ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "read_spreadsheet",
            "description": "Read and summarize a spreadsheet workbook. Returns metadata and optionally preview data for specified sheets.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "include_sheets": {
                        "type": "array",
                        "description": "Optional list of sheet names to include. If omitted and include_all=False, includes all sheets.",
                        "items": {"type": "string"}
                    },
                    "include_all": {
                        "type": "boolean",
                        "description": "If True, include all sheets. Default: True",
                        "default": True
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "Maximum number of rows to read per sheet for preview. Default: 10",
                        "default": 10
                    },
                    "max_cols": {
                        "type": "integer",
                        "description": "Maximum number of columns to read per sheet for preview. Default: 10",
                        "default": 10
                    }
                },
                "required": ["path"]
            }
        }
    })
    async def read_spreadsheet(
        self,
        path: str,
        include_sheets: Optional[List[str]] = None,
        include_all: bool = True,
        max_rows: int = 10,
        max_cols: int = 10
    ) -> ToolResult:
        """Read and summarize spreadsheet"""
        try:
            await self._ensure_sandbox()
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Determine which sheets to include
            if include_sheets:
                sheets_to_read = [s for s in include_sheets if s in workbook.sheetnames]
                if len(sheets_to_read) != len(include_sheets):
                    missing = [s for s in include_sheets if s not in workbook.sheetnames]
                    return self.fail_response(f"Sheet(s) not found: {', '.join(missing)}. Available: {', '.join(workbook.sheetnames)}")
            elif include_all:
                sheets_to_read = workbook.sheetnames
            else:
                sheets_to_read = workbook.sheetnames
            
            # Build metadata and preview
            metadata = {
                "file_path": full_path,
                "sheet_names": workbook.sheetnames,
                "total_sheets": len(workbook.sheetnames)
            }
            
            preview_data = {}
            for sheet_name in sheets_to_read:
                worksheet = workbook[sheet_name]
                preview = self._get_preview_data(worksheet, max_rows=max_rows, max_cols=max_cols)
                preview_data[sheet_name] = {
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "preview": preview
                }
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "metadata": metadata,
                "preview_data": preview_data
            })
            
        except Exception as e:
            logger.error(f"Error reading spreadsheet: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to read spreadsheet: {str(e)}")
    
    # ==================== FUNCTION 9: set_formula ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "set_formula",
            "description": "Set a formula in a cell. The formula string must start with '='. Note: openpyxl stores the formula but does not compute results - Excel will compute when opened.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "cell": {
                        "type": "string",
                        "description": "Cell reference in A1 format (e.g., 'A1', 'B5')"
                    },
                    "formula": {
                        "type": "string",
                        "description": "Formula string starting with '=' (e.g., '=SUM(A1:A10)', '=B2*2')"
                    }
                },
                "required": ["path", "sheet_name", "cell", "formula"]
            }
        }
    })
    async def set_formula(
        self,
        path: str,
        sheet_name: str,
        cell: str,
        formula: str
    ) -> ToolResult:
        """Set a formula in a cell"""
        try:
            await self._ensure_sandbox()
            
            # Validate cell reference
            if not self._validate_cell_reference(cell):
                return self.fail_response(f"Invalid cell reference: {cell}. Expected format: A1")
            
            # Validate formula
            if not formula.startswith('='):
                return self.fail_response(f"Formula must start with '='. Got: {formula}")
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            
            # Set formula
            cell_obj = worksheet[cell.upper()]
            cell_obj.value = formula
            
            # Save back
            await self._save_workbook_to_sandbox(workbook, full_path)
            
            logger.info(f"Set formula in cell {cell} in {full_path}")
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "cell": cell.upper(),
                "formula": formula
            })
            
        except Exception as e:
            logger.error(f"Error setting formula: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to set formula: {str(e)}")
    
    # ==================== merge_range ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "merge_range",
            "description": "Merge a range of cells into one. The top-left cell keeps its value; others are merged into it. Use A1:C1 format for the range.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to existing workbook file"},
                    "sheet_name": {"type": "string", "description": "Name of the sheet"},
                    "range": {"type": "string", "description": "Cell range to merge in A1:C3 format (e.g. A1:N1)"}
                },
                "required": ["path", "sheet_name", "range"]
            }
        }
    })
    async def merge_range(self, path: str, sheet_name: str, range: str) -> ToolResult:
        """Merge a range of cells into one."""
        try:
            await self._ensure_sandbox()
            if not self._validate_range(range):
                return self.fail_response(f"Invalid range format: {range}. Expected format: A1:C3")
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            workbook = await self._load_workbook_from_sandbox(full_path)
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            worksheet.merge_cells(range)
            await self._save_workbook_to_sandbox(workbook, full_path)
            logger.info(f"Merged range {range} in {full_path}")
            return self.success_response({"ok": True, "file_path": full_path, "merged_range": range})
        except Exception as e:
            logger.error(f"Error merging range: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to merge range: {str(e)}")

    # ==================== unmerge_range ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "unmerge_range",
            "description": "Unmerge a previously merged range. Pass the exact range that was merged (e.g. A1:N1). After unmerge, only the top-left cell keeps the value.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to existing workbook file"},
                    "sheet_name": {"type": "string", "description": "Name of the sheet"},
                    "range": {"type": "string", "description": "Merged range to unmerge in A1:C3 format (e.g. A1:N1)"}
                },
                "required": ["path", "sheet_name", "range"]
            }
        }
    })
    async def unmerge_range(self, path: str, sheet_name: str, range: str) -> ToolResult:
        """Unmerge a range of cells."""
        try:
            await self._ensure_sandbox()
            if not self._validate_range(range):
                return self.fail_response(f"Invalid range format: {range}. Expected format: A1:C3")
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            workbook = await self._load_workbook_from_sandbox(full_path)
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            # openpyxl unmerge_cells expects the range string; raises if range is not merged
            range_upper = range.upper()
            merged_refs = [str(m).upper() for m in worksheet.merged_cells.ranges]
            if range_upper not in merged_refs:
                return self.fail_response(
                    f"Range {range} is not merged. Merged ranges in this sheet: {merged_refs if merged_refs else 'none'}"
                )
            worksheet.unmerge_cells(range)
            await self._save_workbook_to_sandbox(workbook, full_path)
            logger.info(f"Unmerged range {range} in {full_path}")
            return self.success_response({"ok": True, "file_path": full_path, "unmerged_range": range})
        except Exception as e:
            logger.error(f"Error unmerging range: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to unmerge range: {str(e)}")
    
    # ==================== FUNCTION 10: format_cells ====================
    @openapi_schema({
        "type": "function",
        "function": {
            "name": "format_cells",
            "description": "Apply formatting to cells. Target can be a single cell (A1) or range (A1:C3). Supports number_format, bold, italic, font_size, font_color, fill_color, alignment, wrap_text, and border.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Path to existing workbook file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Name of the sheet"
                    },
                    "target": {
                        "type": "string",
                        "description": "Single cell (A1) or range (A1:C3) to format"
                    },
                    "format": {
                        "type": "object",
                        "description": "Formatting options object",
                        "properties": {
                            "number_format": {"type": "string", "description": "Number format code (e.g., '0.00', 'mm/dd/yyyy', '$#,##0')"},
                            "bold": {"type": "boolean", "description": "Make text bold"},
                            "italic": {"type": "boolean", "description": "Make text italic"},
                            "font_size": {"type": "integer", "description": "Font size"},
                            "font_color": {"type": "string", "description": "Font color in hex format without # (e.g., 'FF0000' for red)"},
                            "fill_color": {"type": "string", "description": "Background color in hex format without # (e.g., 'FFFF00' for yellow)"},
                            "alignment": {"type": "string", "enum": ["left", "center", "right", "top", "bottom", "middle"], "description": "Text alignment"},
                            "wrap_text": {"type": "boolean", "description": "Wrap text in cells"},
                            "border": {"type": "boolean", "description": "Add border around cells"}
                        }
                    }
                },
                "required": ["path", "sheet_name", "target", "format"]
            }
        }
    })
    async def format_cells(
        self,
        path: str,
        sheet_name: str,
        target: str,
        format: Dict[str, Any]
    ) -> ToolResult:
        """Apply formatting to cells"""
        try:
            await self._ensure_sandbox()
            
            # Determine if target is cell or range
            is_range = ':' in target
            if is_range:
                if not self._validate_range(target):
                    return self.fail_response(f"Invalid range format: {target}. Expected format: A1:C3")
                start_row, start_col, end_row, end_col = self._parse_range(target)
            else:
                if not self._validate_cell_reference(target):
                    return self.fail_response(f"Invalid cell reference: {target}. Expected format: A1")
                # Convert single cell to range
                cell_col = column_index_from_string(re.match(r'^([A-Z]+)', target.upper()).group(1))
                cell_row = int(re.match(r'[A-Z]+(\d+)$', target.upper()).group(1))
                start_row, start_col, end_row, end_col = cell_row, cell_col, cell_row, cell_col
            
            # Validate and normalize path (prevents path traversal)
            try:
                cleaned_path, full_path = self._validate_and_normalize_path(path, must_exist=True)
            except ValueError as e:
                return self.fail_response(f"Invalid path: {str(e)}. Paths must be relative to /workspace.")
            
            # Load workbook
            workbook = await self._load_workbook_from_sandbox(full_path)
            
            # Get worksheet
            if sheet_name not in workbook.sheetnames:
                return self.fail_response(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
            
            # Apply formatting
            applied_keys = []
            for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                for cell in row:
                    # Number format
                    if "number_format" in format:
                        cell.number_format = format["number_format"]
                        applied_keys.append("number_format")
                    
                    # Font formatting
                    font_updates = {}
                    if "bold" in format:
                        font_updates["bold"] = format["bold"]
                        applied_keys.append("bold")
                    if "italic" in format:
                        font_updates["italic"] = format["italic"]
                        applied_keys.append("italic")
                    if "font_size" in format:
                        font_updates["size"] = format["font_size"]
                        applied_keys.append("font_size")
                    if "font_color" in format:
                        from openpyxl.styles.colors import Color
                        font_updates["color"] = Color(rgb=format["font_color"])
                        applied_keys.append("font_color")
                    
                    if font_updates:
                        current_font = cell.font if cell.font else Font()
                        font = Font(
                            bold=font_updates.get("bold", current_font.bold),
                            italic=font_updates.get("italic", current_font.italic),
                            size=font_updates.get("size", current_font.size),
                            color=font_updates.get("color", current_font.color)
                        )
                        cell.font = font
                    
                    # Fill color
                    if "fill_color" in format:
                        fill = PatternFill(start_color=format["fill_color"], end_color=format["fill_color"], fill_type="solid")
                        cell.fill = fill
                        applied_keys.append("fill_color")
                    
                    # Alignment: set both horizontal and vertical explicitly so we fully override
                    # existing alignment (otherwise None on one axis can leave Excel showing old value)
                    if "alignment" in format:
                        current = cell.alignment
                        h = getattr(current, "horizontal", None) if current else None
                        v = getattr(current, "vertical", None) if current else None
                        wrap = getattr(current, "wrap_text", False) if current else False
                        want = format["alignment"]
                        if want in ("left", "center", "right"):
                            h = want
                            if v is None:
                                v = "center"
                        elif want in ("top", "bottom", "middle"):
                            v = "center" if want == "middle" else want
                            if h is None:
                                h = "left"
                        else:
                            h = h or "left"
                            v = v or "center"
                        cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
                        applied_keys.append("alignment")
                    
                    # Wrap text
                    if "wrap_text" in format:
                        if cell.alignment:
                            cell.alignment.wrap_text = format["wrap_text"]
                        else:
                            cell.alignment = Alignment(wrap_text=format["wrap_text"])
                        applied_keys.append("wrap_text")
                    
                    # Border
                    if "border" in format and format["border"]:
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border
                        applied_keys.append("border")
            
            # Save back
            await self._save_workbook_to_sandbox(workbook, full_path)
            
            # Remove duplicates from applied_keys
            applied_keys = list(set(applied_keys))
            
            logger.info(f"Formatted {target} in {full_path}")
            
            return self.success_response({
                "ok": True,
                "file_path": full_path,
                "formatted_target": target,
                "applied_keys": applied_keys
            })
            
        except Exception as e:
            logger.error(f"Error formatting cells: {str(e)}", exc_info=True)
            return self.fail_response(f"Failed to format cells: {str(e)}")
=======
            logger.exception("sb_spreadsheet_tool render_xlsx failed")
            return self.fail_response(f"Failed to render workbook: {str(e)}")
>>>>>>> Stashed changes
